import openpyxl
import pandas as pd
import warnings
warnings.filterwarnings('ignore')

FILE_PATH = ""

def clean_excel_sheets(file_path):
    """
    This function cleans and modifies the excel sheets.

    Parameters:
        file_path (str): The path to the excel file.

    Returns:
        None
    """
    # Read the excel file and get the sheets
    file = pd.ExcelFile(file_path)
    sheet_1 = pd.read_excel(file, 'Outgoing - students')
    sheet_2 = pd.read_excel(file, 'MAP Report')

    # Incorrect course codes, however due diligence required after script completed i.e exchange students
    incorrect_course_codes = ['A0501', 'A0502', '9413']

    # Replace the incorrect course codes with the correct codes
    for index, code in enumerate(sheet_2['Course 1 Code']):
        if str(code) in incorrect_course_codes:
            sheet_2['Course 1 Code'].iloc[index] = sheet_2['Course2 Code'].iloc[index]
            sheet_2['Course 1 Title'].iloc[index] = sheet_2['Course 2 Title'].iloc[index]
            sheet_2['Course 1 Campus'].iloc[index] = sheet_2['Course 2 Campus'].iloc[index]
            sheet_2['Course 1 Managing Faculty'].iloc[index] = sheet_2['Course 2 Managing Faculty'].iloc[index]
            sheet_2['Course 1 Second Faculty'].iloc[index] = sheet_2['Course 2 Second Faculty'].iloc[index]
        else:
            continue

    # Capitalize the first letter of the last name, first name and course campus
    sheet_2['Last Name'] = sheet_2['Last Name'].str.title()
    sheet_2['First Name'] = sheet_2['First Name'].str.title()
    sheet_2['Course 1 Campus'] = sheet_2['Course 1 Campus'].str.title()

    # Remove the unwanted statuses
    sheet_2 = sheet_2[sheet_2['Status'] != 'Unsuccessful']
    sheet_2 = sheet_2[sheet_2['Status'] != 'Pending']

    # Convert the Pre-departure Modules Completion column
    sheet_2.loc[sheet_2['Pre-departure Modules Completion'].notna(), 'Pre-departure Modules Completion'] = True
    sheet_2['Pre-departure Modules Completion'] = sheet_2['Pre-departure Modules Completion'].fillna(False)

    # Convert the INTERNATIONAL_STUDENT column
    sheet_2.loc[sheet_2['INTERNATIONAL_STUDENT'] == 'International', 'INTERNATIONAL_STUDENT'] = True
    sheet_2.loc[sheet_2['INTERNATIONAL_STUDENT'] == 'Domestic', 'INTERNATIONAL_STUDENT'] = False

    # Save the excel sheet
    with pd.ExcelWriter('MA Access database outgoing student data import output.xlsx') as file:
        sheet_1.to_excel(file, 'Outgoing - students',index=False)
        sheet_2.to_excel(file, 'MAP Report',index=False)

clean_excel_sheets(FILE_PATH)

# Load workbook
WB_PATH = ""
wb = openpyxl.load_workbook(WB_PATH)

# Index and assign worksheets to variables for faster access
main_ws = wb.worksheets[0]
MAP_export = wb.worksheets[1]

# Convert 'International student?' column to 'TRUE' and 'FALSE'
for cell in MAP_export["AC"]:
    if cell.value is None:
        break
    elif cell.value == "International":
        cell.value = "TRUE"
    elif cell.value == "Domestic":
        cell.value = "FALSE"

def copy_paste_to_template(copy_column, paste_sheet, paste_column):
    """
    Copies values from `copy_column` and pastes them to `paste_sheet` in column `paste_column`.

    Args:
    - copy_column (openpyxl.cell.cell): The source column to copy.
    - paste_sheet (openpyxl.worksheet.worksheet): The destination sheet to paste to.
    - paste_column (int): The destination column to paste to.

    Returns:
    None
    """
    for cell in copy_column:
        paste_sheet.cell(row = cell[0].row, column = paste_column, value = cell[0].value)

def date_copy_paste(copy_column, paste_sheet, paste_column):
    """
    Copies dates from `copy_column` and pastes them to `paste_sheet` in column `paste_column` as strings in the format `dd-mm-yyyy`.

    Args:
    - copy_column (openpyxl.cell.cell): The source column to copy.
    - paste_sheet (openpyxl.worksheet.worksheet): The destination sheet to paste to.
    - paste_column (int): The destination column to paste to.

    Returns:
    None
    """
    for cell in copy_column:
        if cell[0].value is None:
            continue
        else:
            date = cell[0].value
            year = date.strftime("%Y")
            month = date.strftime("%m")
            day = date.strftime("%d")
            date = f"{day}-{month}-{year}"
            paste_sheet.cell(row = cell[0].row, column = paste_column, value = date)



# Copy values from MAP_export to main_ws
copy_paste_to_template(MAP_export['P2:P2000'], main_ws, 4) # STUDENT ID
copy_paste_to_template(MAP_export['A2:A2000'], main_ws, 5) # LAST NAME
copy_paste_to_template(MAP_export['B2:B2000'], main_ws, 6) # FIRST NAME
date_copy_paste(MAP_export['O2:O2000'], main_ws, 7) # DATE OF BIRTH
copy_paste_to_template(MAP_export['N2:N2000'], main_ws, 17) # EMAIL COPY
copy_paste_to_template(MAP_export['U2:U2000'], main_ws, 18) # MANAGING FACULTY COPY
copy_paste_to_template(MAP_export['V2:V2000'], main_ws, 19) # SECONDARY FACULTY COPY
copy_paste_to_template(MAP_export['S2:S2000'], main_ws, 20) # COURSE TITLE COPY
copy_paste_to_template(MAP_export['R2:R2000'], main_ws, 21) # COURSE CODE COPY
copy_paste_to_template(MAP_export['T2:T2000'], main_ws, 22) # CAMPUS COPY
copy_paste_to_template(MAP_export['Q2:Q2000'], main_ws, 24) # UNDERGRAD/POSTGRAD COPY
copy_paste_to_template(MAP_export['AC2:AC2000'], main_ws, 25) # 'INTERNATIONAL STUDENT?' COPY        
copy_paste_to_template(MAP_export['G2:G2000'], main_ws, 116) # STATUS COPY
copy_paste_to_template(MAP_export['AD2:AD2000'], main_ws, 119) # PRE-DEPARTURE COPY
copy_paste_to_template(MAP_export['E2:E2000'], main_ws, 31) # YEAR ABROAD COPY 
date_copy_paste(MAP_export['H2:H2000'], main_ws, 35) # PROGRAM START DATE COPY
date_copy_paste(MAP_export['I2:I2000'], main_ws, 36) # PROGRAM END DATE COPY
copy_paste_to_template(MAP_export['K2:K2000'], main_ws, 40) # COUNTRY OF PROGRAM 
copy_paste_to_template(MAP_export['M2:M2000'], main_ws, 8) # GENDER COPY
copy_paste_to_template(MAP_export['D2:D2000'], main_ws, 32) # PROGRAM TITLE COPY

# Saving file
wb.save(WB_PATH)

# Final touches
    
def final_touches(WB_PATH):
    """
    Fill in missing values in two Excel sheets and write the changes back to the workbook.
    
    Parameters:
        WB_PATH (str): Path to the Excel workbook.
    
    Returns:
        None
    """
    # Open the Excel file
    file = pd.ExcelFile(WB_PATH)
    
    # Read in the two sheets
    sheet_1 = pd.read_excel(file, 'Outgoing - students')
    sheet_2 = pd.read_excel(file, 'MAP Report')
    
    # Fill in missing values
    sheet_1['AcceptedByMA'] = sheet_1['AcceptedByMA'].fillna(True)
    sheet_1['A&CRcvd'] = sheet_1['A&CRcvd'].fillna(True)
    sheet_1['Second Faculty'] = sheet_1['Second Faculty'].fillna('None')
    
    # Write the changes back to the workbook
    with pd.ExcelWriter(WB_PATH) as file:
        sheet_1.to_excel(file, 'Outgoing - students',index=False)
        sheet_2.to_excel(file, 'MAP Report',index=False)
        
final_touches(WB_PATH)